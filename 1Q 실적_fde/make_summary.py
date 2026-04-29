import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 현금주의 Raw CSV 자동 탐색 (현금주의_Raw_*.csv)
_csv_candidates = sorted(glob.glob(os.path.join(BASE_DIR, '현금주의_Raw_*.csv')))
if not _csv_candidates:
    raise FileNotFoundError(f"현금주의_Raw_*.csv 파일을 {BASE_DIR}에서 찾을 수 없습니다.")
df = pd.read_csv(_csv_candidates[-1], encoding='utf-8-sig')

# 카테고리 매핑
def map_category(row):
    cat = str(row['카테고리']).strip()
    item = str(row['상품명']).strip()

    if cat in ['피트니스', '홀리데이']:
        return '01. 피트니스'
    elif cat in ['운동복', '락커']:
        return '02. 옵션상품'
    elif cat == 'PT':
        if item == '안심결제':
            return '05. PT_안심결제'
        elif '크레딧' in item:
            return '04. PT_대관'
        else:
            return '03. PT_연결'
    elif cat == '팀버핏':
        return '07. 팀버핏'
    elif cat == '요가':
        return '09. 요가(PB)'
    elif cat in ['필라테스', '바레/필라테스']:
        return '10. 필라테스(PB)'
    elif cat == '수수료':
        return '11. 수수료'
    elif cat == '골프':
        return '13. 입점_골프'
    elif cat == '테니스':
        return '15. 입점_테니스'
    elif cat == '패시브스트레칭':
        return '16. 입점_스트레칭'
    elif cat == '스쿼시':
        return '17. 입점_스쿼시'
    elif cat == '굿즈':
        return '98. 굿즈'
    elif cat in ['음료', '푸드']:
        return '99. F&B'
    else:
        return f'미분류({cat})'

df['매출분류'] = df.apply(map_category, axis=1)

# 미분류 확인
unclassified = df[df['매출분류'].str.startswith('미분류')]
if len(unclassified) > 0:
    print("미분류 항목:")
    print(unclassified[['카테고리', '상품명', '매출분류']].value_counts().head(20))

# 지점 순서
branch_order = ['역삼ARC', '역삼GFC', '도곡', '논현', '판교', '강변', '가산', '삼성',
                '광화문', '한티', '마곡', '판교벤처타운', 'GFC', '상도', '합정', '신도림']
# 실제 존재하는 지점만
actual_branches = df['지점명'].unique().tolist()
branches = [b for b in branch_order if b in actual_branches]
branches += sorted([b for b in actual_branches if b not in branch_order])

# 카테고리 순서
cat_order = [
    '01. 피트니스', '02. 옵션상품', '03. PT_연결', '04. PT_대관', '05. PT_안심결제',
    '07. 팀버핏', '09. 요가(PB)', '10. 필라테스(PB)', '11. 수수료',
    '13. 입점_골프', '15. 입점_테니스', '16. 입점_스트레칭', '17. 입점_스쿼시',
    '98. 굿즈', '99. F&B'
]
actual_cats = df['매출분류'].unique().tolist()
cats = [c for c in cat_order if c in actual_cats]
cats += sorted([c for c in actual_cats if c not in cat_order])

# 피벗
pivot = df.groupby(['지점명', '매출분류'])['가격_exvat'].sum().unstack(fill_value=0)
pivot = pivot.reindex(index=branches, columns=cats, fill_value=0)
pivot['합계'] = pivot.sum(axis=1)

# Excel 작성
wb = Workbook()
ws = wb.active
ws.title = '1Q 지점별 매출'

# 스타일 정의
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
subheader_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
grand_total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
white_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
black_font = Font(name='맑은 고딕', size=10)
bold_font = Font(name='맑은 고딕', bold=True, size=10)
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
left = Alignment(horizontal='left', vertical='center')

thin = Side(style='thin', color='B8CCE4')
thick = Side(style='medium', color='1F4E79')
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

# 헤더 행
ws.cell(1, 1, '지점명').font = white_font
ws.cell(1, 1).fill = header_fill
ws.cell(1, 1).alignment = center
ws.column_dimensions['A'].width = 14

for ci, cat in enumerate(cats, 2):
    cell = ws.cell(1, ci, cat)
    cell.font = white_font
    cell.fill = subheader_fill
    cell.alignment = center
    ws.column_dimensions[get_column_letter(ci)].width = 16

total_col = len(cats) + 2
ws.cell(1, total_col, '합계').font = white_font
ws.cell(1, total_col).fill = header_fill
ws.cell(1, total_col).alignment = center
ws.column_dimensions[get_column_letter(total_col)].width = 16

ws.row_dimensions[1].height = 30

# 데이터 행
for ri, branch in enumerate(branches, 2):
    ws.cell(ri, 1, branch).font = bold_font
    ws.cell(ri, 1).alignment = left
    ws.cell(ri, 1).border = border_thin

    for ci, cat in enumerate(cats, 2):
        val = int(pivot.loc[branch, cat]) if branch in pivot.index and cat in pivot.columns else 0
        cell = ws.cell(ri, ci, val if val != 0 else '')
        cell.font = black_font
        cell.alignment = right
        cell.border = border_thin
        if val != 0:
            cell.number_format = '#,##0'

    total_val = int(pivot.loc[branch, '합계']) if branch in pivot.index else 0
    tc = ws.cell(ri, total_col, total_val)
    tc.font = bold_font
    tc.alignment = right
    tc.fill = total_fill
    tc.border = border_thin
    tc.number_format = '#,##0'

    ws.row_dimensions[ri].height = 20

# 합계 행
total_row = len(branches) + 2
ws.cell(total_row, 1, '합계').font = bold_font
ws.cell(total_row, 1).fill = grand_total_fill
ws.cell(total_row, 1).alignment = center

for ci, cat in enumerate(cats, 2):
    val = int(pivot[cat].sum())
    cell = ws.cell(total_row, ci, val)
    cell.font = bold_font
    cell.alignment = right
    cell.fill = grand_total_fill
    cell.border = border_thin
    cell.number_format = '#,##0'

grand_total = int(pivot['합계'].sum())
gtc = ws.cell(total_row, total_col, grand_total)
gtc.font = bold_font
gtc.alignment = right
gtc.fill = grand_total_fill
gtc.border = border_thin
gtc.number_format = '#,##0'

ws.row_dimensions[total_row].height = 22
ws.freeze_panes = 'B2'

out_path = r'c:\Users\wlgml\OneDrive\문서\claude\1Q 실적\1Q_지점별매출합계.xlsx'
wb.save(out_path)
print(f"저장 완료: {out_path}")
print(f"\n전체 합계: {grand_total:,}원")
print("\n카테고리별 합계:")
for cat in cats:
    print(f"  {cat}: {int(pivot[cat].sum()):,}")
