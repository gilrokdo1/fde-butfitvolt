import os
import urllib.request, urllib.parse, json
from copy import copy
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE_DIR, '실적 양식.xlsx')
OUTPUT   = os.path.join(BASE_DIR, '1Q 실적_최종본.xlsx')
API_BASE = os.getenv('COST_API_BASE', 'http://localhost:5000')

BRANCHES = [
    '역삼','도곡','신도림','논현','판교','강변','가산',
    '삼성','광화문','한티역','마곡','판교벤처','GFC','합정','상도'
]
MONTHS = ['2026-01', '2026-02', '2026-03']

# ── 헬퍼 ──────────────────────────────────────────────────

def get_pivot(branch):
    url = f'{API_BASE}/api/cost/{urllib.parse.quote(branch)}/pivot'
    with urllib.request.urlopen(url, timeout=60) as r:
        return json.loads(r.read())

def pn(v):
    try: return int(str(v).replace(',','').replace(' ','')) if v else 0
    except: return 0

def build_lookup(rows):
    lu = {}
    for r in rows:
        key = (r['중분류'].strip(), r['소분류'].strip())
        lu[key] = {m: pn(v) for m, v in r['months'].items()}
    return lu

def sum_cats(lookup, mid_filter, small_list, month):
    total = 0
    for (mid, small), months in lookup.items():
        if mid_filter and mid_filter not in mid:
            continue
        if small in small_list:
            total += months.get(month, 0)
    return total

# ── 매핑 정의 (템플릿 행번호 기준, 계약서 삽입 전) ──────────

# leaf rows: row → (중분류 포함필터, [소분류 정확값 리스트])
BASE_LEAF = {
    54: (None, ['1. 임차료']),
    55: (None, ['2. 고정관리비']),
    57: (None, ['3. 전기료']),
    58: (None, ['4. 수도료']),
    59: (None, []),                                     # 가스료 = 0
    60: (None, ['1. 세탁비']),
    63: (None, ['1. 데스크/백오피스']),
    64: (None, ['2. 탈의실/샤워실 소모품']),
    65: (None, ['3. 청소/미화 소모품']),
    66: (None, ['6. 소도구/기구소모품/가구류']),
    67: (None, ['7. TB_소도구/기구소모품/가구류']),
    68: (None, ['8. TB_기타 소모품']),
    # row 69: 계약서 (조건부 삽입)
    70: (None, ['1. 수건/운동복']),
    73: (None, ['1. 수선비']),
    74: (None, []),                                     # 정기점검비 = 0
    75: (None, ['2. TB_수선비']),
    77: (None, ['1-1. BG_상품 광고비(제작/송출)']),
    78: (None, ['1-2. TB_상품 광고비(제작/송출)']),
    79: (None, []),                                     # 기타 = 0
    81: (None, ['3. 렌탈(사업관련)']),
    82: (None, ['4. 툴 사용료']),
    83: (None, ['5. 통신비', '5-1. TB_통신비']),
    84: (None, ['6. 보험료']),
    85: (None, []),                                     # 채용수수료 = 0
    87: (None, ['1. 식대']),
    89: (None, ['팀활동비_BG']),
    90: (None, ['팀활동비_TB']),
    92: (None, ['1. 회원 Rewards']),
    93: (None, ['2. TB_회원 Rewards']),
    95: (None, ['1. 운반비']),
    96: (None, ['3. 여비교통비']),
    97: (None, ['2. TB_스페셜클래스 운영비']),
    100: ('BG_인건비', ['1. CM 기본급']),
    101: ('BG_인건비', ['2. 파트타이머(근로소득)']),
    102: ('BG_인건비', ['1. 사업소득/일용소득/기타소득']),
    103: ('BG_인건비', ['2. 청소/미화 수수료']),
    104: ('BG_인건비', ['5. CM 인센티브']),
    105: ('BG_인건비', ['6. 4대보험료']),
    106: ('BG_인건비', ['7. 퇴직연금']),
    108: ('TB_인건비', ['1. 팀버핏 기본급']),
    109: ('TB_인건비', ['2. TB_프리랜서 강사료']),
    110: ('TB_인건비', ['1. TB_스페셜클래스 강사료']),
    111: ('TB_인건비', ['3. TB_외주업체 수수료']),
    112: ('TB_인건비', ['5. 팀버핏 인센티브']),
    113: ('TB_인건비', ['6. 4대보험료']),
    114: ('TB_인건비', ['7. 퇴직연금']),
}

# 소계 트리: parent → [children]
BASE_SUBS = {
    56:  [57, 58, 59],
    62:  [63, 64, 65, 66, 67, 68],   # 계약서 있으면 69 추가
    69:  [70],
    61:  [62, 69],
    72:  [73, 74],
    71:  [72, 75],
    76:  [77, 78],
    80:  [81, 82, 83, 84, 85],
    88:  [89, 90],
    86:  [87, 88],
    91:  [92, 93],
    94:  [95, 96],
    99:  [100, 101, 102, 103, 104, 105, 106],
    107: [108, 109, 110, 111, 112, 113, 114],
    98:  [99, 107],
    53:  [54, 55, 56, 60, 61, 71, 76, 79, 80, 86, 91, 94, 97],
    52:  [53, 98],
}

def adjust(has_keiyak):
    """계약서 행 삽입 시 행 번호 조정 (69 이상 +1)"""
    if not has_keiyak:
        return dict(BASE_LEAF), {k: list(v) for k, v in BASE_SUBS.items()}

    new_leaf = {}
    for row, v in BASE_LEAF.items():
        new_leaf[row + 1 if row >= 69 else row] = v
    new_leaf[69] = (None, ['4. 계약서'])

    new_subs = {}
    for parent, children in BASE_SUBS.items():
        np = parent + 1 if parent >= 69 else parent
        nc = [c + 1 if c >= 69 else c for c in children]
        if parent == 62:            # 경상소모품에 계약서(69) 추가
            nc.append(69)
        new_subs[np] = nc

    return new_leaf, new_subs

def compute_vals(lookup, leaf_map, sub_map):
    row_vals = {}

    for row, (mid_f, smalls) in leaf_map.items():
        row_vals[row] = {
            m: (sum_cats(lookup, mid_f, smalls, m) if smalls else 0)
            for m in MONTHS
        }

    seen = set()
    def compute(r):
        if r in seen:
            return row_vals.get(r, {m: 0 for m in MONTHS})
        if r in sub_map:
            vals = {m: 0 for m in MONTHS}
            for c in sub_map[r]:
                cv = compute(c)
                for m in MONTHS:
                    vals[m] += cv.get(m, 0)
            row_vals[r] = vals
        seen.add(r)
        return row_vals.get(r, {m: 0 for m in MONTHS})

    for r in sub_map:
        compute(r)
    return row_vals

def copy_style(src, dst):
    if src.has_style:
        dst.font      = copy(src.font)
        dst.border    = copy(src.border)
        dst.fill      = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)

def process_sheet(ws, lookup, branch, has_keiyak):
    # ① 계약서 행 삽입
    if has_keiyak:
        ws.insert_rows(69, 1)
        for col in range(1, 12):
            copy_style(ws.cell(row=68, column=col), ws.cell(row=69, column=col))
        ws.cell(row=69, column=8).value = '계약서'   # H열

    # ② I/J/K 기존 수치 클리어 (1~120행)
    for r in range(1, 121):
        for col in [9, 10, 11]:
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.value = None

    # ③ 행 번호 조정 후 값 계산
    leaf_map, sub_map = adjust(has_keiyak)
    row_vals = compute_vals(lookup, leaf_map, sub_map)

    # ④ 값 기입 (0 포함)
    for row, month_vals in row_vals.items():
        for mi, month in enumerate(MONTHS):
            ws.cell(row=row, column=9 + mi).value = month_vals.get(month, 0)

    # ⑤ "전사" → 지점명 치환
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '전사' in cell.value:
                cell.value = cell.value.replace('전사', branch)

# ── 메인 ──────────────────────────────────────────────────

def main():
    print('템플릿 로드 중...')
    wb = openpyxl.load_workbook(TEMPLATE)
    tmpl = wb.active

    for branch in BRANCHES:
        print(f'  {branch} 처리 중...', end=' ', flush=True)
        try:
            pivot = get_pivot(branch)
        except Exception as e:
            print(f'SKIP ({e})')
            continue

        lookup = build_lookup(pivot.get('rows', []))

        # 계약서 데이터 있는지 확인
        has_keiyak = any(
            small == '4. 계약서' and any(pn(v) for v in months.values())
            for (mid, small), months in lookup.items()
        )

        ws = wb.copy_worksheet(tmpl)
        ws.title = branch
        process_sheet(ws, lookup, branch, has_keiyak)
        print(f'완료{"(계약서 행 추가)" if has_keiyak else ""}')

    wb.remove(tmpl)
    wb.save(OUTPUT)
    print(f'\n저장 완료: {OUTPUT}')

main()
