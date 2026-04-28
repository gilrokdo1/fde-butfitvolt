import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

_csv_candidates = sorted(glob.glob(os.path.join(BASE_DIR, '현금주의_Raw_*.csv')))
if not _csv_candidates:
    raise FileNotFoundError(f"현금주의_Raw_*.csv 파일을 {BASE_DIR}에서 찾을 수 없습니다.")
df = pd.read_csv(_csv_candidates[-1], encoding='utf-8-sig')

# 월 추출
df['월'] = pd.to_datetime(df['결제일']).dt.month.astype(str) + '월'

# 카테고리 매핑
def map_category(row):
    cat = str(row['카테고리']).strip()
    item = str(row['상품명']).strip()
    if cat in ['피트니스', '홀리데이']:
        return '01. 피트니스'
    elif cat in ['운동복', '락커']:
        return '02. 옵션상품'
    elif cat == 'PT':
        if item == '안심결제':
            return '05. PT_안심결제'
        elif '크레딧' in item:
            return '04. PT_대관'
        else:
            return '03. PT_연결'
    elif cat == '팀버핏':
        return '07. 팀버핏'
    elif cat == '요가':
        return '09. 요가(PB)'
    elif cat in ['필라테스', '바레/필라테스']:
        return '10. 필라테스(PB)'
    elif cat == '수수료':
        return '11. 수수료'
    elif cat == '골프':
        return '13. 입점_골프'
    elif cat == '테니스':
        return '15. 입점_테니스'
    elif cat == '패시브스트레칭':
        return '16. 입점_스트레칭'
    elif cat == '스쿼시':
        return '17. 입점_스쿼시'
    elif cat == '굿즈':
        return '98. 굿즈'
    elif cat in ['음료', '푸드']:
        return '99. F&B'
    else:
        return f'미분류({cat})'

df['매출분류'] = df.apply(map_category, axis=1)

# 지점/카테고리 순서
branch_order = ['역삼ARC', '역삼GFC', '도곡', '논현', '판교', '강변', '가산', '삼성',
                '광화문', '한티', '마곡', '판교벤처타운', '합정', '신도림']
actual_branches = df['지점명'].unique().tolist()
branches = [b for b in branch_order if b in actual_branches]
branches += sorted([b for b in actual_branches if b not in branch_order])

cat_order = [
    '01. 피트니스', '02. 옵션상품', '03. PT_연결', '04. PT_대관', '05. PT_안심결제',
    '07. 팀버핏', '09. 요가(PB)', '10. 필라테스(PB)', '11. 수수료',
    '13. 입점_골프', '15. 입점_테니스', '16. 입점_스트레칭', '17. 입점_스쿼시',
    '98. 굿즈', '99. F&B'
]
months = ['1월', '2월', '3월']

# 피벗: 지점 × 카테고리 × 월
pivot = df.groupby(['지점명', '매출분류', '월'])['가격_exvat'].sum()

# 스타일
def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def make_border(color='B8CCE4', style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BRANCH_FILL   = make_fill('1F4E79')
MONTH_FILL    = make_fill('2E75B6')
CAT_FILL      = make_fill('EBF3FB')
TOTAL_ROW_FILL= make_fill('BDD7EE')
TOTAL_COL_FILL= make_fill('D6E4F0')

W_BOLD  = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
BK_BOLD = Font(name='맑은 고딕', bold=True, size=10)
BK_NORM = Font(name='맑은 고딕', size=10)

BORDER  = make_border()
CENTER  = Alignment(horizontal='center', vertical='center')
RIGHT   = Alignment(horizontal='right',  vertical='center')
LEFT    = Alignment(horizontal='left',   vertical='center')

# Workbook
wb = Workbook()
ws = wb.active
ws.title = '지점별 월별 매출'

# 열 너비 고정 (A=지점/카테고리, B=1월, C=2월, D=3월, E=합계)
ws.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[col].width = 16

current_row = 1

for idx, branch in enumerate(branches):
    branch_data = {}
    for cat in cat_order:
        row_vals = {}
        total = 0
        for m in months:
            try:
                val = int(pivot.loc[(branch, cat, m)])
            except KeyError:
                val = 0
            row_vals[m] = val
            total += val
        row_vals['합계'] = total
        branch_data[cat] = row_vals

    # 해당 지점에 데이터 있는 카테고리만
    active_cats = [c for c in cat_order if branch_data[c]['합계'] != 0]
    if not active_cats:
        continue

    # ── 지점명 헤더 ──
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(current_row, 1, f'{idx+1}. {branch}')
    cell.font = W_BOLD
    cell.fill = BRANCH_FILL
    cell.alignment = CENTER
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # ── 월 헤더 ──
    ws.cell(current_row, 1, '카테고리').font = W_BOLD
    ws.cell(current_row, 1).fill = MONTH_FILL
    ws.cell(current_row, 1).alignment = CENTER
    ws.cell(current_row, 1).border = BORDER
    for ci, label in enumerate(months + ['합계'], 2):
        c = ws.cell(current_row, ci, label)
        c.font = W_BOLD
        c.fill = MONTH_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── 카테고리 행 ──
    for cat in active_cats:
        ws.cell(current_row, 1, cat).font = BK_NORM
        ws.cell(current_row, 1).fill = CAT_FILL
        ws.cell(current_row, 1).alignment = LEFT
        ws.cell(current_row, 1).border = BORDER
        for ci, m in enumerate(months + ['합계'], 2):
            val = branch_data[cat][m]
            c = ws.cell(current_row, ci, val if val != 0 else '')
            c.font = BK_BOLD if m == '합계' else BK_NORM
            c.fill = TOTAL_COL_FILL if m == '합계' else PatternFill()
            c.alignment = RIGHT
            c.border = BORDER
            if val != 0:
                c.number_format = '#,##0'
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # ── 합계 행 ──
    ws.cell(current_row, 1, '합계').font = BK_BOLD
    ws.cell(current_row, 1).fill = TOTAL_ROW_FILL
    ws.cell(current_row, 1).alignment = CENTER
    ws.cell(current_row, 1).border = BORDER
    for ci, m in enumerate(months + ['합계'], 2):
        val = sum(branch_data[c][m] for c in active_cats)
        c = ws.cell(current_row, ci, val)
        c.font = BK_BOLD
        c.fill = TOTAL_ROW_FILL
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = '#,##0'
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # 지점 간 간격
    current_row += 2

out_path = r'c:\Users\wlgml\OneDrive\문서\claude\1Q 실적\1Q_지점별월별매출.xlsx'
wb.save(out_path)
print(f"저장 완료: {out_path}")
