import json
import os
import time

import gspread
import openpyxl
from fastapi import APIRouter, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from google.oauth2.service_account import Credentials

router = APIRouter()

# 경로 설정
_ROUTER_DIR = os.path.dirname(os.path.abspath(__file__))
_FDE_DIR    = os.path.dirname(_ROUTER_DIR)
_PARENT_DIR = os.path.dirname(_FDE_DIR)

SERVICE_ACCOUNT_FILE = os.getenv(
    "COST_SERVICE_ACCOUNT_PATH",
    os.path.join(_FDE_DIR, "service_account.json"),
)
COST_OVERRIDES_FILE  = os.path.join(_FDE_DIR, "cost_overrides.json")
STATIC_DIR           = os.path.join(_FDE_DIR, "static")

SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

COST_SHEET_ID = '1EXz7WifOKdilUIPD6bkCSR5jQCSt_zl8KDnk5vo7TkU'

COST_BRANCHES = [
    '역삼', '도곡', '신도림', '논현', '판교', '강변', '가산',
    '삼성', '광화문', '한티역', '마곡', '판교벤처', 'GFC', '합정', '상도',
]

TB_CARDS = {
    '1253', '1774', '1952', '2096', '2359', '2700', '4259', '4499',
    '4500', '9592', '4757', '4812', '4832', '8411', '1240', '5253',
}

CACHE_TTL = 300  # 5분

_cache: dict = {}


def _get_gc():
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        raise FileNotFoundError(
            f"서비스 계정 파일 없음: {SERVICE_ACCOUNT_FILE}\n"
            "EC2에 파일을 배치하거나 COST_SERVICE_ACCOUNT_PATH 환경변수를 설정하세요."
        )
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return gspread.authorize(creds)


def _get_sheet(sheet_id: str, tab: str, refresh: bool = False) -> list:
    key = f"{sheet_id}::{tab}"
    if not refresh and key in _cache:
        fetched_at, data = _cache[key]
        if time.time() - fetched_at < CACHE_TTL:
            return data
    gc = _get_gc()
    ws = gc.open_by_key(sheet_id).worksheet(tab)
    data = ws.get_all_values()
    _cache[key] = (time.time(), data)
    return data


def _load_overrides() -> list:
    if os.path.exists(COST_OVERRIDES_FILE):
        with open(COST_OVERRIDES_FILE, encoding='utf-8') as f:
            return json.load(f)
    return []


def _ovr_match(row_dict: dict, ovr: dict) -> bool:
    if ovr.get('귀속연월') and row_dict.get('귀속연월', '').strip() != ovr['귀속연월']:
        return False
    if ovr.get('의뢰인') and ovr['의뢰인'] not in row_dict.get('의뢰인/수취인', ''):
        return False
    return True


def _parse_teamact_rows(src_branch: str, refresh: bool = False) -> list:
    sheet_name = f'raw_비용DB_{src_branch}'
    try:
        rows = _get_sheet(COST_SHEET_ID, sheet_name, refresh)
        if not rows:
            return []
        headers = rows[0]
        idx_small = headers.index('재분류소항목') if '재분류소항목' in headers else 22
        result = []
        for row in rows[1:]:
            padded = row + [''] * max(0, len(headers) - len(row))
            if '팀활동비' in padded[idx_small]:
                result.append(dict(zip(headers, padded)))
        return result
    except Exception:
        return []


def _get_teamact_split(branch: str, refresh: bool = False) -> dict:
    overrides = _load_overrides()
    excl = [o for o in overrides if o.get('from_branch') == branch]
    incl = [o for o in overrides if o.get('to_branch') == branch]

    def accumulate(result, row_dict, force_tb=None):
        month = row_dict.get('귀속연월', '').strip()
        if not month:
            return
        acct = row_dict.get('계좌', '').strip()
        last4 = acct[-4:] if len(acct) >= 4 else ''
        is_tb = (last4 in TB_CARDS) if force_tb is None else force_tb
        amt_raw = row_dict.get('부가세등 제외', '')
        try:
            amt = int(float(str(amt_raw).replace(',', '').replace(' ', ''))) if str(amt_raw).strip() else 0
        except (ValueError, AttributeError):
            amt = 0
        if month not in result:
            result[month] = {'bg': 0, 'tb': 0}
        if is_tb:
            result[month]['tb'] += amt
        else:
            result[month]['bg'] += amt

    result: dict = {}
    for row_dict in _parse_teamact_rows(branch, refresh):
        if any(_ovr_match(row_dict, o) for o in excl):
            continue
        accumulate(result, row_dict)

    already_fetched: dict = {}
    for ovr in incl:
        from_b = ovr.get('from_branch')
        if not from_b:
            continue
        if from_b not in already_fetched:
            already_fetched[from_b] = _parse_teamact_rows(from_b, refresh)
        for row_dict in already_fetched[from_b]:
            if _ovr_match(row_dict, ovr):
                accumulate(result, row_dict, force_tb=ovr.get('is_tb'))
    return result


def _build_pivot(branch: str, refresh: bool = False):
    sheet_name = f'{branch}_비용 data'
    rows_raw = _get_sheet(COST_SHEET_ID, sheet_name, refresh)
    if not rows_raw:
        return [], []
    headers = rows_raw[0]
    last_big, last_mid = '', ''
    data_rows = []
    for row in rows_raw[1:]:
        if not any(r.strip() for r in row if r):
            continue
        padded = row + [''] * (len(headers) - len(row))
        big = padded[0].strip() or last_big
        mid = padded[1].strip() or last_mid
        small = padded[2].strip() if len(padded) > 2 else ''
        if padded[0].strip():
            last_big = padded[0].strip()
        if padded[1].strip():
            last_mid = padded[1].strip()
        months = {headers[i]: padded[i] for i in range(3, len(headers)) if i < len(padded)}
        data_rows.append({'대분류': big, '중분류': mid, '소분류': small, 'months': months})
    month_cols = headers[3:]
    teamact_split = None
    final_rows = []
    for r in data_rows:
        if '팀활동비' in r['소분류']:
            if teamact_split is None:
                teamact_split = _get_teamact_split(branch, refresh)
            bg_months = {m: str(teamact_split.get(m, {}).get('bg', 0)) if teamact_split.get(m, {}).get('bg', 0) else '' for m in month_cols}
            tb_months = {m: str(teamact_split.get(m, {}).get('tb', 0)) if teamact_split.get(m, {}).get('tb', 0) else '' for m in month_cols}
            final_rows.append({'대분류': r['대분류'], '중분류': r['중분류'], '소분류': '팀활동비_BG', 'months': bg_months})
            final_rows.append({'대분류': r['대분류'], '중분류': r['중분류'], '소분류': '팀활동비_TB', 'months': tb_months})
        else:
            final_rows.append(r)
    return month_cols, final_rows


# ── 엔드포인트 ────────────────────────────────────────────────────

@router.get("/dashboard", response_class=HTMLResponse)
def cost_dashboard():
    path = os.path.join(STATIC_DIR, "cost.html")
    with open(path, encoding='utf-8') as f:
        return HTMLResponse(content=f.read())


@router.get("/branches")
def get_branches():
    return COST_BRANCHES


@router.get("/{branch}/pivot")
def get_pivot(branch: str, refresh: bool = Query(False)):
    try:
        month_cols, rows = _build_pivot(branch, refresh)
        return {"month_cols": month_cols, "rows": rows}
    except Exception as e:
        return {"error": str(e)}


@router.get("/{branch}/detail")
def get_detail(
    branch: str,
    소분류: str = Query(''),
    월: str = Query(''),
    refresh: bool = Query(False),
):
    teamact_filter = None
    search_small = 소분류
    if 소분류.endswith('_BG'):
        search_small = 소분류[:-3]
        teamact_filter = 'bg'
    elif 소분류.endswith('_TB'):
        search_small = 소분류[:-3]
        teamact_filter = 'tb'

    try:
        rows = _get_sheet(COST_SHEET_ID, f'raw_비용DB_{branch}', refresh)
        if not rows:
            return []
        headers = rows[0]
        idx_small = headers.index('재분류소항목') if '재분류소항목' in headers else 22
        idx_month = headers.index('귀속연월')    if '귀속연월'    in headers else 13
        idx_acct  = headers.index('계좌')        if '계좌'        in headers else 1
        overrides = _load_overrides()
        excl_ovrs = [o for o in overrides if o.get('from_branch') == branch] if teamact_filter else []
        incl_ovrs = [o for o in overrides if o.get('to_branch') == branch]   if teamact_filter else []

        result = []
        for row in rows[1:]:
            padded = row + [''] * (len(headers) - len(row))
            raw_small = padded[idx_small].strip()
            if search_small and not (raw_small == search_small or raw_small.endswith(search_small)):
                continue
            if 월 and padded[idx_month].strip() != 월:
                continue
            if teamact_filter:
                row_dict = dict(zip(headers, padded))
                if any(_ovr_match(row_dict, o) for o in excl_ovrs):
                    continue
                last4 = padded[idx_acct].strip()[-4:] if len(padded[idx_acct].strip()) >= 4 else ''
                is_tb = last4 in TB_CARDS
                if teamact_filter == 'tb' and not is_tb:
                    continue
                if teamact_filter == 'bg' and is_tb:
                    continue
            result.append(dict(zip(headers, padded)))

        if teamact_filter:
            already_fetched: dict = {}
            for ovr in incl_ovrs:
                from_b = ovr.get('from_branch')
                if not from_b:
                    continue
                if from_b not in already_fetched:
                    try:
                        src_rows = _get_sheet(COST_SHEET_ID, f'raw_비용DB_{from_b}', refresh)
                        already_fetched[from_b] = (src_rows[0] if src_rows else [], src_rows[1:] if src_rows else [])
                    except Exception:
                        already_fetched[from_b] = ([], [])
                src_headers, src_data = already_fetched[from_b]
                src_idx_small = src_headers.index('재분류소항목') if '재분류소항목' in src_headers else 22
                src_idx_month = src_headers.index('귀속연월')    if '귀속연월'    in src_headers else 13
                src_idx_acct  = src_headers.index('계좌')        if '계좌'        in src_headers else 1
                for row in src_data:
                    padded2 = row + [''] * (len(src_headers) - len(row))
                    raw_small2 = padded2[src_idx_small].strip()
                    if search_small and not (raw_small2 == search_small or raw_small2.endswith(search_small)):
                        continue
                    if 월 and padded2[src_idx_month].strip() != 월:
                        continue
                    row_dict2 = dict(zip(src_headers, padded2))
                    if not _ovr_match(row_dict2, ovr):
                        continue
                    force_tb = ovr.get('is_tb')
                    if teamact_filter == 'tb' and not force_tb:
                        continue
                    if teamact_filter == 'bg' and force_tb:
                        continue
                    result.append(row_dict2)

        return result
    except Exception as e:
        return {"error": str(e)}


@router.get("/download")
def download_excel(year: str = Query('')):
    from io import BytesIO
    from urllib.parse import quote
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = lambda c: PatternFill('solid', fgColor=c)
    HDR = F('2D3748'); BIG = F('EDF2F7')
    SUB = F('E2E8F0'); GTOT = F('BEE3F8')
    thin = Side(style='thin', color='CCCCCC')
    BDR = Border(left=thin, right=thin, top=thin, bottom=thin)

    def pn(v):
        try:
            return int(str(v).replace(',', '').replace(' ', '')) if v else 0
        except (ValueError, AttributeError):
            return 0

    def sc(cell, fill=None, bold=False, color='2D3748', halign='left', num_fmt=None):
        if fill:
            cell.fill = fill
        cell.font = Font(bold=bold, color=color)
        cell.alignment = Alignment(horizontal=halign, vertical='center')
        if num_fmt:
            cell.number_format = num_fmt
        cell.border = BDR

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for branch in COST_BRANCHES:
        try:
            month_cols, data_rows = _build_pivot(branch)
        except Exception:
            continue
        if not month_cols or not data_rows:
            continue
        if year:
            month_cols = [m for m in month_cols if m.startswith(year)]
        if not month_cols:
            continue

        ws = wb.create_sheet(title=branch)
        ws.freeze_panes = 'D2'
        ws.row_dimensions[1].height = 22

        all_cols = ['대분류', '중분류', '소분류'] + month_cols + ['합계']
        for ci, h in enumerate(all_cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = HDR
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = BDR

        big_totals: dict = {}
        grand = {m: 0 for m in month_cols}
        for r in data_rows:
            bg = r['대분류']
            if bg not in big_totals:
                big_totals[bg] = {m: 0 for m in month_cols}
            for m in month_cols:
                v = pn(r['months'].get(m, ''))
                big_totals[bg][m] += v
                grand[m] += v

        def write_summary_row(ri, label, totals_dict, fill, color):
            ws.cell(row=ri, column=1, value=label)
            for ci in range(1, 4):
                sc(ws.cell(row=ri, column=ci), fill=fill, bold=True, color=color, halign='right' if ci == 1 else 'left')
            row_sum = 0
            for mi, m in enumerate(month_cols):
                v = totals_dict[m]
                row_sum += v
                c = ws.cell(row=ri, column=4 + mi, value=v or None)
                sc(c, fill=fill, bold=True, color=color, halign='right', num_fmt='#,##0')
            c2 = ws.cell(row=ri, column=4 + len(month_cols), value=row_sum or None)
            sc(c2, fill=fill, bold=True, color=color, halign='right', num_fmt='#,##0')

        ri = 2
        prev_big = None
        for r in data_rows:
            is_first_big = r['대분류'] != prev_big
            if is_first_big and prev_big:
                write_summary_row(ri, f'▶ {prev_big} 소계', big_totals[prev_big], SUB, '2D3748')
                ri += 1
            row_sum = sum(pn(r['months'].get(m, '')) for m in month_cols)
            ws.cell(row=ri, column=1, value=r['대분류'] if is_first_big else '')
            sc(ws.cell(row=ri, column=1), bold=is_first_big, halign='left')
            ws.cell(row=ri, column=2, value=r['중분류'])
            sc(ws.cell(row=ri, column=2), halign='left')
            ws.cell(row=ri, column=3, value=r['소분류'])
            sc(ws.cell(row=ri, column=3), halign='left')
            for mi, m in enumerate(month_cols):
                v = pn(r['months'].get(m, ''))
                c = ws.cell(row=ri, column=4 + mi, value=v or None)
                sc(c, halign='right', num_fmt='#,##0')
            c2 = ws.cell(row=ri, column=4 + len(month_cols), value=row_sum or None)
            sc(c2, halign='right', num_fmt='#,##0')
            prev_big = r['대분류']
            ri += 1

        if prev_big:
            write_summary_row(ri, f'▶ {prev_big} 소계', big_totals[prev_big], SUB, '2D3748')
            ri += 1
        write_summary_row(ri, '◆ 전체 합계', grand, GTOT, '1A365D')

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 22
        for ci in range(4, 4 + len(month_cols) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = quote('비용_전지점.xlsx')
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
