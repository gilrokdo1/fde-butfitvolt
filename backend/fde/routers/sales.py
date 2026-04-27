"""
매출보고 대시보드 API (Flask server.py → FastAPI 포팅)
prefix: /fde-api/sales
"""
import gzip as _gzip
import json
import os
import re

import gspread
import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from google.oauth2.service_account import Credentials

from utils.db import safe_db

router = APIRouter()


_moneyplus_table_ready = False
_ref_card_table_ready = False

def _ensure_moneyplus_table():
    global _moneyplus_table_ready
    if _moneyplus_table_ready:
        return
    try:
        with safe_db() as (conn, cur):
            cur.execute("""
                CREATE TABLE IF NOT EXISTS jihee_moneyplus (
                    id SERIAL PRIMARY KEY,
                    type VARCHAR(10) NOT NULL,
                    approval_no VARCHAR(50),
                    row_data JSONB NOT NULL,
                    uploaded_at TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_jihee_moneyplus_type ON jihee_moneyplus(type)"
            )
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_jihee_moneyplus_approval
                ON jihee_moneyplus(type, approval_no)
                WHERE approval_no IS NOT NULL AND approval_no != ''
            """)
        _moneyplus_table_ready = True
    except Exception as e:
        print(f"[sales] jihee_moneyplus 테이블 생성 실패: {e}")


def _ensure_ref_card_table():
    global _ref_card_table_ready
    if _ref_card_table_ready:
        return
    try:
        with safe_db() as (conn, cur):
            cur.execute("""
                CREATE TABLE IF NOT EXISTS jihee_ref_card (
                    id SERIAL PRIMARY KEY,
                    지점명 VARCHAR(50) NOT NULL,
                    카드사명 VARCHAR(50),
                    가맹점번호 VARCHAR(50) NOT NULL,
                    비고 VARCHAR(200) DEFAULT ''
                )
            """)
            cur.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_jihee_ref_card_merchant "
                "ON jihee_ref_card(가맹점번호)"
            )
        _ref_card_table_ready = True
        # JSON 파일이 있으면 마이그레이션
        _migrate_ref_card_from_json()
    except Exception as e:
        print(f"[sales] jihee_ref_card 테이블 생성 실패: {e}")


def _migrate_ref_card_from_json():
    if not os.path.exists(REF_CARD_FILE):
        return
    try:
        rows = load_json(REF_CARD_FILE, [])
        if not rows:
            return
        with safe_db() as (conn, cur):
            cur.execute("SELECT COUNT(*) as cnt FROM jihee_ref_card")
            if cur.fetchone()["cnt"] > 0:
                return  # 이미 데이터 있으면 스킵
        for row in rows:
            try:
                with safe_db() as (conn, cur):
                    cur.execute(
                        """INSERT INTO jihee_ref_card(지점명, 카드사명, 가맹점번호, 비고)
                           VALUES (%s,%s,%s,%s)
                           ON CONFLICT (가맹점번호) DO UPDATE
                           SET 지점명=EXCLUDED.지점명, 카드사명=EXCLUDED.카드사명, 비고=EXCLUDED.비고""",
                        (row.get("지점명",""), row.get("카드사명",""),
                         str(row.get("가맹점번호","")).strip(), row.get("비고",""))
                    )
            except Exception:
                pass
        print(f"[sales] ref_card JSON→DB 마이그레이션 완료: {len(rows)}건")
    except Exception as e:
        print(f"[sales] ref_card 마이그레이션 실패: {e}")


def _load_ref_card_from_db() -> list:
    _ensure_ref_card_table()
    with safe_db() as (conn, cur):
        cur.execute("SELECT id, 지점명, 카드사명, 가맹점번호, 비고 FROM jihee_ref_card ORDER BY id")
        return [dict(r) for r in cur.fetchall()]


def _save_ref_card_to_db(rows: list):
    _ensure_ref_card_table()
    with safe_db() as (conn, cur):
        cur.execute("DELETE FROM jihee_ref_card")
    for row in rows:
        try:
            with safe_db() as (conn, cur):
                cur.execute(
                    """INSERT INTO jihee_ref_card(지점명, 카드사명, 가맹점번호, 비고)
                       VALUES (%s,%s,%s,%s)""",
                    (row.get("지점명",""), row.get("카드사명",""),
                     str(row.get("가맹점번호","")).strip(), row.get("비고",""))
                )
        except Exception as e:
            print(f"[sales] ref_card insert 오류: {e}")

# ── 경로 설정 ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 데이터 파일은 백엔드 루트 기준 sales_data/ 폴더에 저장
DATA_DIR = os.path.join(os.path.dirname(BASE_DIR), "sales_data")
os.makedirs(DATA_DIR, exist_ok=True)

SERVICE_ACCOUNT_FILE = os.path.join(DATA_DIR, "service_account.json")
CONFIG_FILE = os.path.join(DATA_DIR, "config.json")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

SALESFILES_DIR = os.path.join(DATA_DIR, "salesfiles")
os.makedirs(SALESFILES_DIR, exist_ok=True)

CARD_DATA_FILE = os.path.join(UPLOADS_DIR, "card_data.json")
CASH_DATA_FILE = os.path.join(UPLOADS_DIR, "cash_data.json")
OTHER_DATA_FILE = os.path.join(UPLOADS_DIR, "other_data.json")
PORTONE_DATA_FILE = os.path.join(UPLOADS_DIR, "portone_data.json")
REF_CARD_FILE = os.path.join(UPLOADS_DIR, "ref_card.json")
REF_CASH_FILE = os.path.join(UPLOADS_DIR, "ref_cash.json")
REF_OTHER_ROWS_FILE = os.path.join(UPLOADS_DIR, "ref_other_rows.json")
ARCHIVE_FILE = os.path.join(UPLOADS_DIR, "archive.json")

# 임대인 정산 raw 엑셀 (업로드 시 저장 위치)
LOCAL_EXCEL = os.path.join(DATA_DIR, "지점별 임대인_장기분할공급 정산 raw.xlsx")

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
_sheets_cache: dict = {}

# ── 헬퍼 ──────────────────────────────────────────────────────────────────

def _default_config() -> dict:
    return {
        "main_sheet_id": "",
        "branches": [
            {"id": "역삼본관", "name": "역삼ARC", "sales_sheet_id": "", "color": "#3182F6"},
            {"id": "역삼별관", "name": "역삼빗썸", "sales_sheet_id": "", "color": "#1B64DA"},
            {"id": "도곡", "name": "도곡", "sales_sheet_id": "", "color": "#00B493"},
            {"id": "신도림", "name": "신도림", "sales_sheet_id": "", "color": "#FF6B00"},
            {"id": "논현", "name": "논현", "sales_sheet_id": "", "color": "#8B5CF6"},
            {"id": "판교", "name": "판교", "sales_sheet_id": "", "color": "#EC4899"},
            {"id": "강변", "name": "강변", "sales_sheet_id": "", "color": "#06B6D4"},
            {"id": "가산", "name": "가산", "sales_sheet_id": "", "color": "#F59E0B"},
            {"id": "삼성", "name": "삼성", "sales_sheet_id": "", "color": "#10B981"},
            {"id": "광화문", "name": "광화문", "sales_sheet_id": "", "color": "#EF4444"},
            {"id": "한티", "name": "한티", "sales_sheet_id": "", "color": "#6366F1"},
            {"id": "마곡", "name": "마곡", "sales_sheet_id": "", "color": "#14B8A6"},
            {"id": "판교벤처", "name": "판교벤처", "sales_sheet_id": "", "color": "#F97316"},
            {"id": "GFC", "name": "GFC", "sales_sheet_id": "", "color": "#84CC16"},
        ],
    }


def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return json.load(f)
    return _default_config()


def load_json(path: str, default=None):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return default if default is not None else []


def save_json(path: str, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_gc():
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return gspread.authorize(creds)


def get_sheet_data(sheet_id: str, tab_name: str, use_cache: bool = True):
    cache_key = f"{sheet_id}::{tab_name}"
    if use_cache and cache_key in _sheets_cache:
        return _sheets_cache[cache_key]
    gc = get_gc()
    ws = gc.open_by_key(sheet_id).worksheet(tab_name)
    data = ws.get_all_values()
    _sheets_cache[cache_key] = data
    return data


# ── salesfile 캐시 ────────────────────────────────────────────────────────
_salesfile_cache: dict = {}
_salesfile_resp: dict = {}
_salesfile_rows_c: dict = {}


def _get_salesfile_mtime(kind: str) -> float:
    mtime = 0.0
    for f in os.listdir(SALESFILES_DIR):
        m = re.search(r"매출_상세내역\((.+?)\)", f)
        if not m or not f.lower().endswith(".xlsx"):
            continue
        label = m.group(1)
        is_cash = "현금" in label
        if (kind == "cash") != is_cash:
            continue
        try:
            mtime = max(mtime, os.path.getmtime(os.path.join(SALESFILES_DIR, f)))
        except OSError:
            pass
    return mtime


def read_excel_safe(path: str, header_none: bool = False):
    if not header_none:
        try:
            mt = os.path.getmtime(path)
            if path in _salesfile_cache and _salesfile_cache[path][0] == mt:
                return _salesfile_cache[path][1].copy()
        except OSError:
            pass

    with pd.ExcelFile(path) as xf:
        raw = pd.read_excel(xf, xf.sheet_names[0], header=None, dtype=str)

    if header_none:
        return raw

    hi = None
    for i, row in raw.iterrows():
        if str(row.iloc[0]).strip() == "번호":
            hi = i
            break
    if hi is None:
        return None

    df = raw.iloc[hi + 1 :].copy()
    df.columns = raw.iloc[hi].tolist()
    df = df.reset_index(drop=True)

    try:
        mt = os.path.getmtime(path)
        _salesfile_cache[path] = (mt, df)
    except OSError:
        pass
    return df.copy()


def _get_merchant_map() -> dict:
    try:
        ref = _load_ref_card_from_db()
    except Exception:
        ref = load_json(REF_CARD_FILE, [])
    return {str(r.get("가맹점번호", "")).strip(): r.get("지점명", "") for r in ref if r.get("가맹점번호")}


def read_salesfiles_as_card():
    other_rows_mtime = os.path.getmtime(REF_OTHER_ROWS_FILE) if os.path.exists(REF_OTHER_ROWS_FILE) else 0
    xlsx_mtime = _get_salesfile_mtime("card")
    cache_key = (xlsx_mtime, other_rows_mtime)
    if "card" in _salesfile_rows_c and _salesfile_rows_c["card"][0] == cache_key:
        return _salesfile_rows_c["card"][1]

    merchant_map = _get_merchant_map()
    other_row_map = load_json(REF_OTHER_ROWS_FILE, {})
    result = []
    for f in os.listdir(SALESFILES_DIR):
        m = re.search(r"매출_상세내역\((.+?)\)", f)
        if not m or not f.lower().endswith(".xlsx"):
            continue
        label = m.group(1)
        if "현금" in label:
            continue
        is_other = label == "기타"
        path = os.path.join(SALESFILES_DIR, f)
        try:
            df = read_excel_safe(path)
            if df is None:
                continue
            df = df.dropna(how="all").fillna("")
            df.columns = [str(c).strip() for c in df.columns]
            if is_other:
                if "승인번호" in df.columns:
                    df["지점구분"] = df["승인번호"].apply(lambda x: other_row_map.get(str(x).strip(), ""))
                else:
                    df["지점구분"] = ""
            else:
                if "가맹점번호" in df.columns:
                    df["지점구분"] = df["가맹점번호"].apply(lambda x: merchant_map.get(str(x).strip(), ""))
                else:
                    df["지점구분"] = ""
            result.extend(df.to_dict(orient="records"))
        except Exception:
            continue
    _salesfile_rows_c["card"] = (cache_key, result)
    return result


def read_salesfiles_as_cash():
    ref_mtime = os.path.getmtime(REF_CASH_FILE) if os.path.exists(REF_CASH_FILE) else 0
    xlsx_mtime = _get_salesfile_mtime("cash")
    cache_key = (xlsx_mtime, ref_mtime)
    if "cash" in _salesfile_rows_c and _salesfile_rows_c["cash"][0] == cache_key:
        return _salesfile_rows_c["cash"][1]

    ref = load_json(REF_CASH_FILE, [])
    terminal_map = {str(r.get("단말기번호", "")).strip(): r.get("지점구분", "") for r in ref if r.get("단말기번호")}

    result = []
    for f in os.listdir(SALESFILES_DIR):
        m = re.search(r"매출_상세내역\((.+?)\)", f)
        if not m or not f.lower().endswith(".xlsx"):
            continue
        label = m.group(1)
        if "현금" not in label:
            continue
        path = os.path.join(SALESFILES_DIR, f)
        try:
            df = read_excel_safe(path)
            if df is None:
                continue
            df = df.dropna(how="all").fillna("")
            df.columns = [str(c).strip() for c in df.columns]
            if "단말기번호" in df.columns:
                df["지점구분"] = df["단말기번호"].apply(lambda x: terminal_map.get(str(x).strip(), ""))
            else:
                df["지점구분"] = ""
            result.extend(df.to_dict(orient="records"))
        except Exception:
            continue
    _salesfile_rows_c["cash"] = (cache_key, result)
    return result


# ── 설정 ──────────────────────────────────────────────────────────────────

@router.get("/config")
def api_config_get():
    return load_config()


@router.post("/config")
async def api_config_save(request: Request):
    data = await request.json()
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _sheets_cache.clear()
    return {"ok": True}


# ── 월별 data (구글 시트) ─────────────────────────────────────────────────

@router.get("/data/monthly")
def api_monthly(branch: str = ""):
    cfg = load_config()
    if not cfg.get("main_sheet_id"):
        return []
    try:
        rows = get_sheet_data(cfg["main_sheet_id"], "월별data")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    if not rows:
        return []
    headers = rows[1] if len(rows) > 1 else rows[0]
    try:
        branch_col = headers.index("지점명_분류")
    except ValueError:
        try:
            branch_col = headers.index("지점명")
        except ValueError:
            branch_col = 14
    result = []
    for row in rows[2:]:
        if not any(row):
            continue
        if branch and (len(row) <= branch_col or row[branch_col] != branch):
            continue
        result.append(dict(zip(headers, row)))
    return result


# ── 카드 데이터 ───────────────────────────────────────────────────────────

def _load_moneyplus(type_: str) -> list:
    _ensure_moneyplus_table()
    with safe_db() as (conn, cur):
        cur.execute("SELECT row_data FROM jihee_moneyplus WHERE type=%s ORDER BY id", (type_,))
        return [dict(r["row_data"]) for r in cur.fetchall()]

def _upsert_moneyplus(type_: str, rows: list) -> int:
    _ensure_moneyplus_table()
    added = 0
    for row in rows:
        key = (row.get("승인번호") or "").strip() or None
        try:
            with safe_db() as (conn, cur):
                if key:
                    cur.execute(
                        """INSERT INTO jihee_moneyplus(type, approval_no, row_data)
                           VALUES (%s,%s,%s)
                           ON CONFLICT (type, approval_no)
                           WHERE approval_no IS NOT NULL AND approval_no != ''
                           DO NOTHING""",
                        (type_, key, json.dumps(row, ensure_ascii=False))
                    )
                else:
                    cur.execute(
                        "INSERT INTO jihee_moneyplus(type, row_data) VALUES (%s,%s)",
                        (type_, json.dumps(row, ensure_ascii=False))
                    )
                if cur.rowcount:
                    added += 1
        except Exception as e:
            print(f"[sales] upsert 오류: {e}")
    return added

def _delete_moneyplus(type_: str):
    with safe_db() as (conn, cur):
        cur.execute("DELETE FROM jihee_moneyplus WHERE type=%s", (type_,))

def _count_moneyplus(type_: str) -> int:
    with safe_db() as (conn, cur):
        cur.execute("SELECT COUNT(*) as cnt FROM jihee_moneyplus WHERE type=%s", (type_,))
        return cur.fetchone()["cnt"]


@router.get("/data/card")
def api_card(branch: str = ""):
    data = _load_moneyplus("card") + load_json(OTHER_DATA_FILE, []) + read_salesfiles_as_card()
    if not data:
        return []
    merchant_map = _get_merchant_map()
    result = []
    for row in data:
        if not row.get("지점구분"):
            mn = str(row.get("가맹점번호", "")).strip()
            row["지점구분"] = merchant_map.get(mn, "")
        if branch and row.get("지점구분", "") != branch:
            continue
        result.append(row)
    return result


@router.post("/upload/card")
async def upload_card(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="데이터 없음")
    added = _upsert_moneyplus("card", data)
    total = _count_moneyplus("card")
    return {"ok": True, "added": added, "total": total}


@router.post("/upload/card/replace")
async def upload_card_replace(request: Request):
    data = await request.json() or []
    _delete_moneyplus("card")
    _upsert_moneyplus("card", data)
    total = _count_moneyplus("card")
    return {"ok": True, "total": total}


@router.delete("/upload/card")
def delete_card():
    _delete_moneyplus("card")
    return {"ok": True}


# ── 기타 데이터 ───────────────────────────────────────────────────────────

@router.get("/data/other")
def api_other():
    return load_json(OTHER_DATA_FILE, [])


@router.post("/upload/other")
async def upload_other(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="데이터 없음")
    existing = load_json(OTHER_DATA_FILE, [])
    existing_keys = {r.get("승인번호", "") for r in existing if r.get("승인번호")}
    added = 0
    for row in data:
        key = row.get("승인번호", "")
        if key and key in existing_keys:
            continue
        existing.append(row)
        if key:
            existing_keys.add(key)
        added += 1
    save_json(OTHER_DATA_FILE, existing)
    return {"ok": True, "added": added, "total": len(existing)}


@router.delete("/upload/other")
def delete_other():
    save_json(OTHER_DATA_FILE, [])
    return {"ok": True}


# ── 현금 데이터 ───────────────────────────────────────────────────────────

@router.get("/data/cash")
def api_cash(branch: str = ""):
    data = _load_moneyplus("cash") + read_salesfiles_as_cash()
    if not data:
        return []
    ref = load_json(REF_CASH_FILE, [])
    terminal_map = {str(r.get("단말기번호", "")).strip(): r.get("지점구분", "") for r in ref if r.get("단말기번호")}
    result = []
    for row in data:
        if not row.get("지점구분"):
            tn = str(row.get("단말기번호", "")).strip()
            row["지점구분"] = terminal_map.get(tn, "")
        if branch and row.get("지점구분", "") != branch:
            continue
        result.append(row)
    return result


@router.post("/upload/cash")
async def upload_cash(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="데이터 없음")
    added = _upsert_moneyplus("cash", data)
    total = _count_moneyplus("cash")
    return {"ok": True, "added": added, "total": total}


@router.post("/upload/cash/replace")
async def upload_cash_replace(request: Request):
    data = await request.json() or []
    _delete_moneyplus("cash")
    _upsert_moneyplus("cash", data)
    total = _count_moneyplus("cash")
    return {"ok": True, "total": total}


@router.delete("/upload/cash")
def delete_cash():
    _delete_moneyplus("cash")
    return {"ok": True}


# ── ref 값 ────────────────────────────────────────────────────────────────

@router.get("/ref/card")
def ref_card_get():
    try:
        return _load_ref_card_from_db()
    except Exception:
        return load_json(REF_CARD_FILE, [])


@router.post("/ref/card")
async def ref_card_save(request: Request):
    rows = await request.json() or []
    _save_ref_card_to_db(rows)
    save_json(REF_CARD_FILE, rows)  # JSON 백업 유지
    _salesfile_rows_c.pop("card", None)
    return {"ok": True}


@router.get("/ref/cash")
def ref_cash_get():
    return load_json(REF_CASH_FILE, [])


@router.post("/ref/cash")
async def ref_cash_save(request: Request):
    save_json(REF_CASH_FILE, await request.json() or [])
    return {"ok": True}


@router.get("/ref/other-rows")
def ref_other_rows_get():
    return load_json(REF_OTHER_ROWS_FILE, {})


@router.post("/ref/other-rows")
async def ref_other_rows_save(request: Request):
    body = await request.json() or {}
    data = load_json(REF_OTHER_ROWS_FILE, {})
    data.update(body)
    save_json(REF_OTHER_ROWS_FILE, data)
    _salesfile_resp.clear()
    _salesfile_rows_c.pop("card", None)
    return {"ok": True}


# ── ref 단일 행 ───────────────────────────────────────────────────────────

@router.post("/ref/{rtype}/row")
async def ref_row_add(rtype: str, request: Request):
    row = await request.json()
    if rtype == "card":
        try:
            with safe_db() as (conn, cur):
                cur.execute(
                    """INSERT INTO jihee_ref_card(지점명, 카드사명, 가맹점번호, 비고)
                       VALUES (%s,%s,%s,%s)
                       ON CONFLICT (가맹점번호) DO UPDATE
                       SET 지점명=EXCLUDED.지점명, 카드사명=EXCLUDED.카드사명, 비고=EXCLUDED.비고""",
                    (row.get("지점명",""), row.get("카드사명",""),
                     str(row.get("가맹점번호","")).strip(), row.get("비고",""))
                )
            _salesfile_rows_c.pop("card", None)
        except Exception as e:
            print(f"[sales] ref_row_add card 오류: {e}")
    else:
        file = REF_CASH_FILE
        data = load_json(file, [])
        data.append(row)
        save_json(file, data)
    return {"ok": True}


@router.put("/ref/{rtype}/row/{idx}")
async def ref_row_update(rtype: str, idx: int, request: Request):
    row = await request.json()
    if rtype == "card":
        try:
            with safe_db() as (conn, cur):
                cur.execute(
                    "SELECT id FROM jihee_ref_card ORDER BY id OFFSET %s LIMIT 1", (idx,)
                )
                rec = cur.fetchone()
                if rec:
                    cur.execute(
                        """UPDATE jihee_ref_card SET 지점명=%s, 카드사명=%s, 가맹점번호=%s, 비고=%s
                           WHERE id=%s""",
                        (row.get("지점명",""), row.get("카드사명",""),
                         str(row.get("가맹점번호","")).strip(), row.get("비고",""), rec["id"])
                    )
            _salesfile_rows_c.pop("card", None)
        except Exception as e:
            print(f"[sales] ref_row_update card 오류: {e}")
    else:
        file = REF_CASH_FILE
        data = load_json(file, [])
        if 0 <= idx < len(data):
            data[idx] = row
            save_json(file, data)
    return {"ok": True}


@router.delete("/ref/{rtype}/row/{idx}")
def ref_row_delete(rtype: str, idx: int):
    if rtype == "card":
        try:
            with safe_db() as (conn, cur):
                cur.execute(
                    "SELECT id FROM jihee_ref_card ORDER BY id OFFSET %s LIMIT 1", (idx,)
                )
                rec = cur.fetchone()
                if rec:
                    cur.execute("DELETE FROM jihee_ref_card WHERE id=%s", (rec["id"],))
            _salesfile_rows_c.pop("card", None)
        except Exception as e:
            print(f"[sales] ref_row_delete card 오류: {e}")
    else:
        file = REF_CASH_FILE
        data = load_json(file, [])
        if 0 <= idx < len(data):
            data.pop(idx)
            save_json(file, data)
    return {"ok": True}


# ── ref 초기화 ────────────────────────────────────────────────────────────

@router.post("/ref/init-from-excel")
def ref_init_excel():
    if not os.path.exists(LOCAL_EXCEL):
        raise HTTPException(status_code=404, detail="엑셀 파일을 찾을 수 없습니다")
    try:
        wb = openpyxl.load_workbook(LOCAL_EXCEL, read_only=True, data_only=True)
        ws_card = wb["DB_지점별 가맹점 번호"]
        card_ref = []
        for row in list(ws_card.iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            card_ref.append({
                "지점명": str(row[0] or "").strip(),
                "카드사명": str(row[1] or "").strip(),
                "가맹점번호": str(row[2] or "").strip(),
                "비고": str(row[3] or "").strip(),
            })
        save_json(REF_CARD_FILE, card_ref)
        _save_ref_card_to_db(card_ref)
        ws_cash = wb["DB_현금 단말기"]
        cash_ref = []
        for row in list(ws_cash.iter_rows(values_only=True))[1:]:
            if not row[1]:
                continue
            cash_ref.append({
                "단말기번호": str(row[1] or "").strip(),
                "단말기명": str(row[2] or "").strip(),
                "지점구분": str(row[3] or "").strip(),
            })
        save_json(REF_CASH_FILE, cash_ref)
        wb.close()
        return {"ok": True, "card": len(card_ref), "cash": len(cash_ref)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/ref/init")
def ref_init():
    cfg = load_config()
    try:
        rows = get_sheet_data(cfg["main_sheet_id"], "DB_지점별 가맹점 번호", use_cache=False)
        card_ref = []
        if rows:
            headers = rows[0]
            card_ref = [dict(zip(headers, r)) for r in rows[1:] if any(r) and r[0]]
            save_json(REF_CARD_FILE, card_ref)
            _save_ref_card_to_db(card_ref)
        rows2 = get_sheet_data(cfg["main_sheet_id"], "DB_현금 단말기", use_cache=False)
        cash_ref = []
        if rows2:
            headers2 = rows2[0]
            cash_ref = [dict(zip(headers2, r)) for r in rows2[1:] if any(r)]
            save_json(REF_CASH_FILE, cash_ref)
        return {"ok": True, "card": len(card_ref), "cash": len(cash_ref)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── 매출_상세내역 salesfile 목록/데이터 ──────────────────────────────────

@router.get("/salesfiles")
def salesfiles_list():
    files = []
    for f in sorted(os.listdir(SALESFILES_DIR)):
        m = re.search(r"매출_상세내역\((.+?)\)", f)
        if m and f.lower().endswith(".xlsx"):
            files.append({"label": m.group(1), "filename": f})
    return files


@router.get("/salesfiles/{label}")
def salesfiles_data(label: str):
    for f in os.listdir(SALESFILES_DIR):
        m = re.search(r"매출_상세내역\((.+?)\)", f)
        if m and m.group(1) == label and f.lower().endswith(".xlsx"):
            path = os.path.join(SALESFILES_DIR, f)
            try:
                is_cash = "현금" in label
                is_other = label == "기타"
                ref_file = REF_CASH_FILE if is_cash else REF_CARD_FILE
                other_rows_mtime = os.path.getmtime(REF_OTHER_ROWS_FILE) if is_other and os.path.exists(REF_OTHER_ROWS_FILE) else 0

                try:
                    mtime_xlsx = os.path.getmtime(path)
                    mtime_ref = os.path.getmtime(ref_file) if os.path.exists(ref_file) else 0
                    cache_key = (label, mtime_xlsx, mtime_ref, other_rows_mtime)
                    if label in _salesfile_resp and _salesfile_resp[label][0] == cache_key:
                        gz_bytes = _salesfile_resp[label][1]
                        return Response(
                            content=gz_bytes,
                            media_type="application/json; charset=utf-8",
                            headers={"Content-Encoding": "gzip"},
                        )
                except OSError:
                    cache_key = None

                df = read_excel_safe(path)
                if df is None:
                    raise HTTPException(status_code=500, detail="헤더를 찾을 수 없습니다")
                df = df.dropna(how="all")
                df.columns = [str(c).strip() for c in df.columns]
                df = df.fillna("")

                date_col = ""
                for c in df.columns:
                    for v in df[c].dropna().head(5):
                        if re.match(r"^\d{8}$", str(v)) or re.match(r"^\d{4}-\d{2}-\d{2}", str(v)):
                            date_col = c
                            break
                    if date_col:
                        break

                name_col = ""
                for c in df.columns:
                    for v in df[c].dropna().head(10):
                        sv = str(v).strip()
                        if len(sv) > 5 and not re.match(r"^[\d\-\*\s/]+$", sv):
                            name_col = c
                            break
                    if name_col:
                        break

                if is_other:
                    key_col = "승인번호"
                    row_map = load_json(REF_OTHER_ROWS_FILE, {})
                    if key_col in df.columns:
                        df.insert(0, "지점", df[key_col].apply(lambda x: row_map.get(str(x).strip(), "")))
                    else:
                        df.insert(0, "지점", "")
                else:
                    key_col = "단말기번호" if is_cash else "가맹점번호"
                    val_col = "지점구분" if is_cash else "지점명"
                    if is_cash:
                        ref = load_json(ref_file, [])
                    else:
                        ref = _load_ref_card_from_db()
                    ref_map = {str(r.get(key_col, "")).strip(): r.get(val_col, "") for r in ref if r.get(key_col)}
                    if key_col in df.columns:
                        df.insert(0, "지점", df[key_col].apply(lambda x: ref_map.get(str(x).strip(), "")))
                    else:
                        df.insert(0, "지점", "")

                cols = list(df.columns)
                result = df.to_dict(orient="records")
                resp_data = {
                    "columns": cols,
                    "rows": result,
                    "date_col": date_col,
                    "name_col": name_col,
                    "key_col": key_col if key_col in cols else "",
                    "label": label,
                }

                resp_json = json.dumps(resp_data, ensure_ascii=False).encode("utf-8")
                gz_bytes = _gzip.compress(resp_json, compresslevel=6)
                if cache_key:
                    _salesfile_resp[label] = (cache_key, gz_bytes)

                return Response(
                    content=gz_bytes,
                    media_type="application/json; charset=utf-8",
                    headers={"Content-Encoding": "gzip"},
                )
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=500, detail=str(e))

    raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")


# ── 포트원 ────────────────────────────────────────────────────────────────

@router.get("/data/portone")
def api_portone(branch: str = ""):
    result = []
    local = load_json(PORTONE_DATA_FILE, [])
    for row in local:
        if branch:
            order_name = str(row.get("주문명", "") or row.get("상품명", "") or "")
            if branch not in order_name:
                continue
        result.append(row)
    try:
        cfg = load_config()
        if cfg.get("main_sheet_id"):
            rows = get_sheet_data(cfg["main_sheet_id"], "온라인_포트원")
            if rows:
                headers = rows[0]
                for row in rows[1:]:
                    if not any(row):
                        continue
                    if branch:
                        order_name = row[5] if len(row) > 5 else ""
                        if branch not in order_name:
                            continue
                    result.append(dict(zip(headers, row)))
    except Exception:
        pass
    return result


@router.post("/upload/portone")
async def upload_portone(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="데이터 없음")
    existing = load_json(PORTONE_DATA_FILE, [])
    dup_keys: set = set()
    for r in existing:
        k = r.get("imp_uid") or r.get("결제번호") or r.get("merchant_uid") or r.get("주문번호") or ""
        if k:
            dup_keys.add(str(k).strip())
    added = 0
    for row in data:
        k = row.get("imp_uid") or row.get("결제번호") or row.get("merchant_uid") or row.get("주문번호") or ""
        k = str(k).strip()
        if k and k in dup_keys:
            continue
        existing.append(row)
        if k:
            dup_keys.add(k)
        added += 1
    save_json(PORTONE_DATA_FILE, existing)
    return {"ok": True, "added": added, "total": len(existing)}


@router.delete("/upload/portone")
def delete_portone():
    save_json(PORTONE_DATA_FILE, [])
    return {"ok": True}


@router.get("/portone/count")
def portone_count():
    return {"count": len(load_json(PORTONE_DATA_FILE, []))}


# ── 지점별 매출보고 시트 탭 ───────────────────────────────────────────────

@router.get("/branch-sheet/{sheet_id}/tabs")
def api_branch_tabs(sheet_id: str):
    try:
        gc = get_gc()
        ss = gc.open_by_key(sheet_id)
        tabs = [ws.title for ws in ss.worksheets()]
        return tabs
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/branch-sheet/{sheet_id}/{tab_name:path}")
def api_branch_sheet(sheet_id: str, tab_name: str, refresh: str = "false"):
    try:
        rows = get_sheet_data(sheet_id, tab_name, use_cache=(refresh != "true"))
        return rows
    except gspread.exceptions.WorksheetNotFound:
        raise HTTPException(status_code=404, detail=f"탭을 찾을 수 없습니다: {tab_name}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── 업로드 이력 ───────────────────────────────────────────────────────────

@router.get("/archive")
def archive_get():
    return load_json(ARCHIVE_FILE, [])


@router.post("/archive")
async def archive_add(request: Request):
    entry = await request.json()
    archive = load_json(ARCHIVE_FILE, [])
    archive.insert(0, entry)
    save_json(ARCHIVE_FILE, archive)
    return {"ok": True}


@router.delete("/archive/{idx}")
def archive_delete(idx: int):
    archive = load_json(ARCHIVE_FILE, [])
    if 0 <= idx < len(archive):
        archive.pop(idx)
        save_json(ARCHIVE_FILE, archive)
    return {"ok": True}


# ── 캐시 초기화 ───────────────────────────────────────────────────────────

@router.post("/cache/clear")
def api_cache_clear():
    _sheets_cache.clear()
    return {"ok": True}
